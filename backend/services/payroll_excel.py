import io
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.services.attendance import format_day_label
from backend.services.payroll_calc import days_in_month, weekday_labels

if TYPE_CHECKING:
    from backend.models import PayrollEmployee, PayrollModule


def _employee_row_dict(employee: "PayrollEmployee", employee_response) -> dict:
    att = employee_response.attendance or {}
    return {
        "serial_no": employee_response.serial_no,
        "name": employee_response.name,
        "designation": employee_response.designation or "",
        "attendance": att,
        "total_ot_hours": employee_response.total_ot_hours,
        "ot_rate": employee_response.ot_rate,
        "ot_amount": employee_response.ot_amount,
        "total_days": employee_response.total_days,
        "advance": employee_response.advance,
        "wage": employee_response.wage,
        "monthly_salary": employee_response.monthly_salary,
        "food": employee_response.food or "",
        "final_payment": employee_response.final_payment,
        "remarks": employee_response.remarks or "",
    }


def build_payroll_workbook(
    module: "PayrollModule",
    employees: list["PayrollEmployee"],
    employee_responses: list,
) -> bytes:
    day_count = days_in_month(module.year, module.month)
    weekdays = weekday_labels(module.year, module.month)
    sunday_days = {day for day in range(1, day_count + 1) if weekdays[day - 1].startswith("Sun")}

    wb = Workbook()
    ws = wb.active
    ws.title = "Salaries"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E8F5")
    sunday_fill = PatternFill("solid", fgColor="FFF3CD")
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center")

    fixed_headers = ["S.No", "Name", "Designation"]
    tail_headers = [
        "OT Hrs",
        "OT Rate",
        "OT Amt",
        "Total Days",
        "Advance",
        "Wage",
        "Salary/Month",
        "Food",
        "Final Payment",
        "Remarks",
    ]
    last_col = len(fixed_headers) + day_count + len(tail_headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=module.title)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_row = 2
    weekday_row = 3
    col = 1
    for label in fixed_headers:
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
        ws.cell(row=weekday_row, column=col, value="").border = border
        col += 1

    for day in range(1, day_count + 1):
        fill = sunday_fill if day in sunday_days else header_fill
        head = ws.cell(row=header_row, column=col, value=day)
        head.font = header_font
        head.fill = fill
        head.border = border
        head.alignment = center
        wd = ws.cell(row=weekday_row, column=col, value=weekdays[day - 1])
        wd.fill = fill
        wd.border = border
        wd.alignment = center
        col += 1

    for label in tail_headers:
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
        ws.cell(row=weekday_row, column=col, value="").border = border
        col += 1

    data_start = 4
    for row_idx, (emp, resp) in enumerate(zip(employees, employee_responses, strict=False)):
        row_num = data_start + row_idx
        data = _employee_row_dict(emp, resp)
        col = 1
        ws.cell(row=row_num, column=col, value=data["serial_no"]).border = border
        col += 1
        ws.cell(row=row_num, column=col, value=data["name"]).border = border
        col += 1
        ws.cell(row=row_num, column=col, value=data["designation"]).border = border
        col += 1
        for day in range(1, day_count + 1):
            raw_day = data["attendance"].get(str(day), "")
            cell = ws.cell(row=row_num, column=col, value=format_day_label(raw_day) or "·")
            cell.border = border
            cell.alignment = center
            if day in sunday_days:
                cell.fill = sunday_fill
            col += 1
        for key in (
            "total_ot_hours",
            "ot_rate",
            "ot_amount",
            "total_days",
            "advance",
            "wage",
            "monthly_salary",
            "food",
            "final_payment",
            "remarks",
        ):
            cell = ws.cell(row=row_num, column=col, value=data[key])
            cell.border = border
            if key in (
                "total_ot_hours",
                "ot_rate",
                "ot_amount",
                "advance",
                "wage",
                "monthly_salary",
                "food",
                "final_payment",
            ):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            col += 1

    total_row = data_start + len(employees)
    total_payment = sum(r.final_payment for r in employee_responses)
    ws.cell(row=total_row, column=1, value="")
    ws.cell(row=total_row, column=2, value="")
    total_label = ws.cell(row=total_row, column=3, value="Total")
    total_label.font = Font(bold=True)
    total_label.border = border
    pay_col = len(fixed_headers) + day_count + 8
    total_cell = ws.cell(row=total_row, column=pay_col, value=total_payment)
    total_cell.font = Font(bold=True)
    total_cell.border = border

    ws.freeze_panes = "D4"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    for day_col in range(4, 4 + day_count):
        ws.column_dimensions[get_column_letter(day_col)].width = 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
